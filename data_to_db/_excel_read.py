"""审查入库结果 - 读取 Excel 原始数据用于对比"""
import openpyxl
import os

data_dir = r"c:\Users\31039\ExecelDataDeal\data\纵向多子表"
files = [f for f in os.listdir(data_dir) if f.endswith('.xlsx') or f.endswith('.xls')]
files.sort()

for f in files:
    path = os.path.join(data_dir, f)
    print(f"\n{'='*80}")
    print(f"文件: {f}")
    print(f"{'='*80}")
    wb = openpyxl.load_workbook(path, data_only=True)
    for sname in wb.sheetnames:
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        print(f"\n  Sheet: {sname} ({ws.max_row} rows x {ws.max_column} cols)")
        # 打印前20行
        for i, row in enumerate(rows[:20]):
            print(f"    Row {i+1}: {list(row)}")
        if len(rows) > 20:
            print(f"    ... ({len(rows)-20} more rows)")
    wb.close()
