"""端到端 IO 测试：临时合成 xlsx，验证 excel_preview / excel_reader / strategy_standard。"""

import os
import shutil
import tempfile
import unittest

import openpyxl

import _bootstrap  # noqa: F401


def _make_xlsx_basic(path: str):
    """生成一个含 2 个 sheet 的 xlsx：
    - Sheet 'Data'：标准面板（年份, 地区, 产量），3 行数据
    - Sheet '空表'：空数据
    """
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Data"
    ws1.append(["年份", "地区", "产量"])
    ws1.append([2020, "北京", 1234])
    ws1.append([2021, "上海", 5678])
    ws1.append([2022, "广州", 999])

    ws2 = wb.create_sheet("空表")

    wb.save(path)


def _make_xlsx_with_merged(path: str):
    """生成含合并单元格的 xlsx：第 1 行是标题（合并 A1:C1），第 2 行是表头，3+ 行是数据。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MergedTitle"
    ws["A1"] = "测试报表标题"
    ws.merge_cells("A1:C1")
    ws.append(["年份", "地区", "产量"])  # 第 2 行
    ws.append([2020, "北京", 1])
    ws.append([2021, "上海", 2])
    wb.save(path)


class TestExcelPreviewAndReader(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.tmpdir = tempfile.mkdtemp(prefix="datadeal_test_")
        cls.basic_path = os.path.join(cls.tmpdir, "basic.xlsx")
        cls.merged_path = os.path.join(cls.tmpdir, "merged.xlsx")
        _make_xlsx_basic(cls.basic_path)
        _make_xlsx_with_merged(cls.merged_path)

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.tmpdir, ignore_errors=True)

    def test_list_sheets(self):
        from services.excel_preview import list_sheets
        info = list_sheets(self.basic_path)
        self.assertIn("Data", info["sheet_names"])
        self.assertIn("空表", info["sheet_names"])
        self.assertFalse(info["is_xls"])
        self.assertNotIn("error", info)

    def test_preview_run(self):
        from services.excel_preview import run as preview_run
        info = preview_run(self.basic_path, sheet_index=0, preview_rows=10)
        self.assertEqual(info["sheet_name"], "Data")
        self.assertGreaterEqual(info["max_row"], 4)
        self.assertEqual(info["max_col"], 3)
        # 第一行应该是表头
        self.assertEqual(info["preview_data"][0][:3], ["年份", "地区", "产量"])

    def test_read_sheet_basic(self):
        from services.excel_reader import read_sheet
        data, _ = read_sheet(self.basic_path, "Data", read_border=False)
        self.assertGreaterEqual(len(data), 4)
        self.assertEqual(data[0][:3], ["年份", "地区", "产量"])
        # 数据列允许 int 或 str（openpyxl 默认保留类型）
        self.assertEqual(str(data[1][0]), "2020")
        self.assertEqual(data[1][1], "北京")

    def test_read_sheet_with_merged(self):
        """合并单元格应被 read_sheet 填充展开。"""
        from services.excel_reader import read_sheet
        data, _ = read_sheet(self.merged_path, "MergedTitle", read_border=False)
        # 第一行 A1:C1 合并 → 三列都该是同一值
        self.assertEqual(data[0][0], "测试报表标题")
        self.assertEqual(data[0][1], "测试报表标题")
        self.assertEqual(data[0][2], "测试报表标题")


class TestPreprocessSheet(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.tmpdir = tempfile.mkdtemp(prefix="datadeal_pre_")
        cls.path = os.path.join(cls.tmpdir, "with_footnote.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "S1"
        ws.append([None, None])
        ws.append(["年份", "GDP"])
        ws.append([2020, 1000.5])
        ws.append([2021, 1100.7])
        ws.append([None, None])
        ws.append(["资料来源：国家统计局", None])
        ws.append([None, None])
        wb.save(cls.path)

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.tmpdir, ignore_errors=True)

    def test_basic(self):
        from services.table_layout import preprocess_sheet
        info = preprocess_sheet(self.path, "S1", read_border=False)
        # 首行非空数据应是 ["年份", "GDP"]
        self.assertEqual(info["data"][0][0], "年份")
        # 不包括脚注行；不包括尾部空行
        self.assertEqual(len(info["data"]), 3)
        for row in info["data"]:
            self.assertFalse(all(v is None for v in row),
                             "preprocess 后不应残留全空行")
        # footnote 至少识别到 1 行
        self.assertGreaterEqual(info["footnote_trim"], 1)


class TestStrategyStandard(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.tmpdir = tempfile.mkdtemp(prefix="datadeal_test_std_")
        cls.basic_path = os.path.join(cls.tmpdir, "basic.xlsx")
        _make_xlsx_basic(cls.basic_path)

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.tmpdir, ignore_errors=True)

    def test_run_returns_columns_and_rows(self):
        from strategies.strategy_standard import run as std_run
        result = std_run(self.basic_path, "Data", "ods_test", column_names=None)
        self.assertEqual(result["columns"], ["年份", "地区", "产量"])
        self.assertEqual(len(result["rows"]), 3)
        self.assertEqual(result["rows"][0][1], "北京")

    def test_column_names_override(self):
        from strategies.strategy_standard import run as std_run
        result = std_run(self.basic_path, "Data", "ods_test",
                         column_names=["year", "region", "output"])
        self.assertEqual(result["columns"], ["year", "region", "output"])
        # 列数不一致时应忽略覆盖
        result2 = std_run(self.basic_path, "Data", "ods_test",
                          column_names=["year", "region"])  # 少一列
        self.assertEqual(result2["columns"], ["年份", "地区", "产量"])


if __name__ == "__main__":
    unittest.main()
