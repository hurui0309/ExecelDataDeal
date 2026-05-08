"""测试 services/excel_reader.read_sheet 的 mtime-aware 缓存。"""

import os
import shutil
import tempfile
import time
import unittest

import openpyxl

import _bootstrap  # noqa: F401


class TestReadSheetCache(unittest.TestCase):
    """每个 test 用独立 tmpdir，避免缓存/文件状态串扰。"""

    def setUp(self):
        from services.excel_reader import clear_cache
        clear_cache()
        self.tmpdir = tempfile.mkdtemp(prefix="datadeal_cache_")
        self.path = os.path.join(self.tmpdir, "tiny.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "S1"
        ws.append(["年份", "GDP"])
        ws.append([2020, 1000])
        ws.append([2021, 1100])
        wb.save(self.path)

    def tearDown(self):
        from services.excel_reader import clear_cache
        clear_cache()
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_returns_independent_lists(self):
        """缓存命中时返回的 data 应当与缓存独立，调用方修改不应污染下次读取。"""
        from services.excel_reader import read_sheet
        d1, _ = read_sheet(self.path, "S1")
        d1[0][0] = "MUTATED"
        d2, _ = read_sheet(self.path, "S1")
        self.assertNotEqual(d2[0][0], "MUTATED")
        self.assertEqual(d2[0][0], "年份")

    def test_cache_hit_faster(self):
        """第二次调用比第一次快得多（缓存命中）。"""
        from services.excel_reader import read_sheet
        t0 = time.perf_counter()
        read_sheet(self.path, "S1")
        first = time.perf_counter() - t0
        t0 = time.perf_counter()
        read_sheet(self.path, "S1")
        second = time.perf_counter() - t0
        self.assertLess(second, first, f"second={second:.4f} should be < first={first:.4f}")

    def test_cache_invalidate_on_mtime_change(self):
        """改 mtime 后，缓存应失效，重新读取得到新数据。"""
        from services.excel_reader import read_sheet
        d1, _ = read_sheet(self.path, "S1")
        self.assertEqual(d1[0][0], "年份")

        wb = openpyxl.load_workbook(self.path)
        ws = wb.active
        ws.cell(row=1, column=1, value="UPDATED")
        wb.save(self.path)
        new_time = os.path.getmtime(self.path) + 2
        os.utime(self.path, (new_time, new_time))

        d2, _ = read_sheet(self.path, "S1")
        self.assertEqual(d2[0][0], "UPDATED")


if __name__ == "__main__":
    unittest.main()
