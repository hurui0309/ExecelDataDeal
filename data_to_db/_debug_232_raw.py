"""直接用 openpyxl 读取232城市表，检查单元格内换行"""
import openpyxl

EXCEL = r'c:\Users\31039\ExecelDataDeal\data\纵向多子表\232个城市主要经济指标(一~四).xlsx'
wb = openpyxl.load_workbook(EXCEL, data_only=True)
ws = wb['1-4']

# 查找含换行的单元格
print("=== 含换行符的单元格 ===")
for row in ws.iter_rows(min_row=1, max_row=40, max_col=7):
    for cell in row:
        if cell.value is not None and '\n' in str(cell.value):
            print(f"  Cell({cell.row},{cell.column}): {repr(str(cell.value))}")

# 打印前40行原始数据
print("\n=== 原始数据前40行 ===")
for r in range(1, 41):
    vals = []
    for c in range(1, 8):
        v = ws.cell(r, c).value
        vals.append(v)
    # 检查是否几乎空行（只有1-2个非空值）
    non_empty = [(c, repr(str(v))) for c, v in enumerate(vals) if v is not None and str(v).strip()]
    marker = ""
    if 0 < len(non_empty) <= 2:
        marker = " <<SPARSE>>"
    display = [str(v)[:25] if v is not None else '' for v in vals]
    print(f"Row {r:2d}: {display}{marker}")

wb.close()
