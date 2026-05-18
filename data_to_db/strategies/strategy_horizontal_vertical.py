"""策略: strategy_horizontal_vertical — 横向分区 + 纵向分页重复表头混合表

适用场景：
  同一 Sheet 内，左右两段数据并排（中间以纵向点划线 dashDot 分隔），
  且全表被"分页重复表头"切成若干相同结构的块（每块约 40 行数据，
  块间以若干空行 + 重复标题行隔开）。

检测条件（全部依赖框线，无文本启发式）：
  1. 某列的 left=dashDot 存在（纵向分割线），且占全部行比例 >= 0.01
     → 该列为横向分割列（split_col），左段=[0:split_col]，右段=[split_col:]
  2. 行级 bottom_solid ratio>=0.7 的行存在
     → detect_header_end_by_border() 能找到表头范围

解析流程：
  Step 1  读框线 → 定 split_col + header 行数（header_rows = header_end+1）
  Step 2  找"分页块"边界：col[split_col] 的 left=dashDot 的行（每块最后一行）
  Step 3  从第一块提取合并表头（左/右各自独立，调用 _merge_multi_row_header）
  Step 4  逐块提取数据行，分左/右两段，跳过表头行和全空行
  Step 5  合并所有块的左段 → left_table，右段 → right_table
  Step 6  返回 subtables=[left_table, right_table]
"""

from __future__ import annotations

import logging

from services.excel_reader import read_sheet
from services.excel_utils import is_empty_row, rename_id_col
from services.border_info import (
    read_border_info,
    detect_header_end_by_border,
    find_dash_left_cols,
)
from services.table_layout import (
    clean_cell,
    trim_trailing_empty_cols,
    truncate_footnotes,
    find_data_start_row,
)

logger = logging.getLogger("datadeal")


DESCRIPTION = (
    "横向+纵向混合表。特征：左右两段数据并排（中间以纵向点划线分隔），同时全表被"
    "\u201c分页重复表头\u201d切成若干相同结构的块（每块约 N 行数据，块间有空行+重复标题行）。"
    "关键框线信号：某列 left=dashDot（纵向分割线）同时存在水平实线表头边框。"
    "典型场景：统计年鉴中每页左右各印一张同结构表的排版。"
)


# ──────────────────────────────────────────────────────────────────────────────
# 公共入口
# ──────────────────────────────────────────────────────────────────────────────

def run(file_path: str, sheet_name: str, table_name: str,
        column_names: list = None, params: dict = None,
        llm_client=None) -> dict:
    """解析横向+纵向混合表。

    params（可选，通常由 pre_classify_by_border 自动填入）：
        split_col     : 横向分割列索引（0-based），不提供则自动从框线检测
        header_end    : 第一块表头结束行（0-based），不提供则自动检测
    """
    params = params or {}
    p_split_col  = params.get("split_col")
    p_header_end = params.get("header_end")

    # ── 读数据 + 框线 ──────────────────────────────────────────────
    data, _ = read_sheet(file_path, sheet_name, read_border=False)
    if not data:
        return {"subtables": [], "original_row_count": 0}

    border_info = read_border_info(file_path, sheet_name)
    rows_info   = border_info["rows"] if border_info else []
    cols_info   = border_info["cols"] if border_info else []

    # ── Step 1: 确定分割列 ──────────────────────────────────────────
    split_col = _resolve_split_col(p_split_col, cols_info, data)
    if split_col is None or split_col >= len(data[0] if data else []):
        logger.warning("strategy_horizontal_vertical: 无法确定分割列，退出")
        return {"subtables": [], "original_row_count": 0}

    # ── Step 2: 找所有分页块的边界行（需先于表头检测，因为要限制检测范围） ──
    block_last_rows = _find_block_boundaries(
        file_path, sheet_name, cols_info, split_col, len(data)
    )
    if not block_last_rows:
        logger.warning("strategy_horizontal_vertical: 未找到分页块边界，退出")
        return {"subtables": [], "original_row_count": 0}

    # ── Step 3: 在第一个块范围内确定表头行数 ────────────────────────
    first_block_end_row = block_last_rows[0]
    first_block_rows_info = rows_info[:first_block_end_row + 1] if rows_info else []
    header_end = _resolve_header_end(
        p_header_end, first_block_rows_info, data[:first_block_end_row + 1], split_col
    )
    header_rows = header_end + 1   # 表头占行数（0-based → 共 header_rows 行）

    logger.info(
        f"  [HV] split_col={split_col}, header_rows={header_rows}, "
        f"blocks={len(block_last_rows)}"
    )

    # ── Step 4: 从第一块提取合并表头 ────────────────────────────────
    first_block_start = 0
    first_block_end   = block_last_rows[0]
    first_block_data  = data[first_block_start : first_block_end + 1]

    left_header_raw  = first_block_data[:header_rows]
    right_header_raw = [row[split_col:] for row in left_header_raw]
    left_header_raw  = [row[:split_col] for row in left_header_raw]

    left_columns  = _build_columns(left_header_raw,  0,            header_end)
    right_columns = _build_columns(right_header_raw, 0,            header_end)

    # ── Step 5: 逐块提取数据行 ──────────────────────────────────────
    left_rows_all:  list[list] = []
    right_rows_all: list[list] = []

    block_starts = [0] + [r + 1 for r in block_last_rows[:-1]]
    for b_start, b_last in zip(block_starts, block_last_rows):
        block_data = data[b_start : b_last + 1]
        _extract_block_rows(
            block_data, header_rows, split_col,
            len(left_columns), len(right_columns),
            left_rows_all, right_rows_all,
        )

    logger.info(
        f"  [HV] left_rows={len(left_rows_all)}, right_rows={len(right_rows_all)}"
    )

    subtables = []
    if left_rows_all:
        subtables.append({
            "columns": left_columns,
            "rows":    left_rows_all,
            "label":   "left",
        })
    if right_rows_all:
        subtables.append({
            "columns": right_columns,
            "rows":    right_rows_all,
            "label":   "right",
        })

    total = sum(len(s["rows"]) for s in subtables)
    return {"subtables": subtables, "original_row_count": total}


# ──────────────────────────────────────────────────────────────────────────────
# 内部辅助
# ──────────────────────────────────────────────────────────────────────────────

def _resolve_split_col(p_split_col, cols_info: list, data: list) -> int | None:
    """确定横向分割列。优先用 params，其次检测 left_dash。"""
    if p_split_col is not None:
        return int(p_split_col)
    if cols_info:
        dash_left = find_dash_left_cols(cols_info, min_ratio=0.01)
        if dash_left:
            return dash_left[0]
    # 最后兜底：尝试用空列检测（来自 horizontal_split 的逻辑）
    if data:
        from strategies.strategy_horizontal_split import _detect_split_col
        return _detect_split_col(data)
    return None


def _resolve_header_end(p_header_end, rows_info: list, data: list,
                         split_col: int) -> int:
    """确定第一块的 header_end（0-based，即表头最后一行索引）。"""
    if p_header_end is not None:
        return int(p_header_end)
    if rows_info:
        he = detect_header_end_by_border(rows_info)
        if he is not None:
            return he
    # 内容回退：找左段的 data_start，header_end = data_start - 1
    left_data = [row[:split_col] for row in data[:min(20, len(data))]]
    ds = find_data_start_row(left_data)
    return max(0, ds - 1) if ds > 0 else 0


def _find_block_boundaries(file_path: str, sheet_name,
                            cols_info: list, split_col: int,
                            total_rows: int) -> list[int]:
    """返回每个分页块最后一行的行索引（0-based）。

    主信号：openpyxl 直接读 col[split_col] 的 left border，
    找 dashDot 的行 → 每块的最后一行。
    回退：若 cols_info 不可用，按等间距估算（不推荐，仅兜底）。
    """
    from services.excel_utils import is_xls_file
    import openpyxl

    if is_xls_file(file_path):
        # xls：用 cols_info 的 left 信息（xlrd 读取的 left_dash 行位置无法按行枚举，
        # 改用"行 bottom_solid 后跟连续空行"的模式回退）
        return _find_blocks_by_empty_rows(
            file_path, sheet_name, split_col, total_rows
        )

    # xlsx：直接逐行读 col[split_col] 的 left border
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.worksheets[0] if isinstance(sheet_name, int) else wb[sheet_name]
        col_excel = split_col + 1   # 1-based
        boundary_rows = []
        for r in range(1, ws.max_row + 1):
            cell = ws.cell(r, col_excel)
            lb = cell.border.left.style if cell.border and cell.border.left else None
            if lb in {
                'dashDot', 'dashDotDot', 'dashed', 'dotted', 'slantDashDot',
                'mediumDashed', 'mediumDashDot', 'mediumDashDotDot',
            }:
                boundary_rows.append(r - 1)   # 0-based
        wb.close()
        if boundary_rows:
            return boundary_rows
    except Exception as e:
        logger.warning(f"_find_block_boundaries xlsx read failed: {e}")

    return _find_blocks_by_empty_rows(file_path, sheet_name, split_col, total_rows)


def _find_blocks_by_empty_rows(file_path: str, sheet_name,
                                 split_col: int, total_rows: int) -> list[int]:
    """回退策略：把连续空行之前的最后一个有数据行作为块边界。

    适用于框线信息不可用时（xls 格式 / 框线读取失败）。
    """
    data, _ = read_sheet(file_path, sheet_name, read_border=False)
    if not data:
        return []
    boundaries = []
    in_empty = False
    last_data_row = -1
    for i, row in enumerate(data):
        left_nonempty = any(
            v is not None and str(v).strip()
            for v in row[:split_col]
        )
        right_nonempty = any(
            v is not None and str(v).strip()
            for v in row[split_col:]
        )
        if left_nonempty or right_nonempty:
            if in_empty and last_data_row >= 0:
                boundaries.append(last_data_row)
            in_empty = False
            last_data_row = i
        else:
            in_empty = True
    if last_data_row >= 0 and (not boundaries or boundaries[-1] != last_data_row):
        boundaries.append(last_data_row)
    return boundaries


def _build_columns(header_rows_data: list, header_start: int, header_end: int) -> list[str]:
    """从多行表头数据合并出列名列表（不依赖 pymysql/mysql_writer）。

    简单合并规则：
    - 单行表头：直接取列值
    - 多行表头：同列各行非空值用 "_" 拼接（父级 + 子级）
    """
    if not header_rows_data:
        return []

    n_cols = max((len(r) for r in header_rows_data), default=0)
    cols = []
    for ci in range(n_cols):
        parts = []
        prev = None
        for row in header_rows_data:
            v = row[ci] if ci < len(row) else None
            if v is not None:
                s = str(v).strip()
                if s and s != prev:
                    parts.append(s)
                    prev = s
        cols.append("_".join(parts) if parts else f"col{ci}")

    cols = _make_unique_columns(cols)
    cols = [rename_id_col(c) for c in cols]
    return cols


def _make_unique_columns(columns: list) -> list[str]:
    """简版列名唯一化（与 mysql_writer.make_unique_columns 等效，不依赖 pymysql）。"""
    result = []
    seen: dict[str, int] = {}
    for col in columns:
        base = str(col) if col is not None else "col"
        if base not in seen:
            seen[base] = 0
            result.append(base)
        else:
            seen[base] += 1
            candidate = f"{base}_{seen[base]}"
            # 避免与已有列名碰撞
            while candidate in seen:
                seen[base] += 1
                candidate = f"{base}_{seen[base]}"
            seen[candidate] = 0
            result.append(candidate)
    return result


def _extract_block_rows(
    block_data: list,
    header_rows: int,
    split_col: int,
    n_left_cols: int,
    n_right_cols: int,
    left_out: list,
    right_out: list,
) -> None:
    """从一个分页块中提取数据行，追加到 left_out / right_out。

    - 跳过前 header_rows 行（表头）
    - 跳过全空行
    - 左段 / 右段各自清洗 + 补齐列数
    """
    for i, row in enumerate(block_data):
        if i < header_rows:
            continue
        if is_empty_row(row):
            continue
        # 左段
        left_vals = [clean_cell(v) for v in row[:split_col]]
        while len(left_vals) < n_left_cols:
            left_vals.append(None)
        left_out.append(left_vals[:n_left_cols])
        # 右段
        right_vals = [clean_cell(v) for v in row[split_col:]]
        while len(right_vals) < n_right_cols:
            right_vals.append(None)
        right_out.append(right_vals[:n_right_cols])
