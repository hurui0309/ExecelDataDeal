"""扫描232城市表全部稀疏行（续行），统计问题规模"""
import openpyxl

EXCEL = r'c:\Users\31039\ExecelDataDeal\data\纵向多子表\232个城市主要经济指标(一~四).xlsx'
wb = openpyxl.load_workbook(EXCEL, data_only=True)
ws = wb['1-4']

sparse_rows = []
data_row_count = 0
for r in range(1, ws.max_row + 1):
    vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
    non_empty = [(c-1, str(v).strip()) for c, v in enumerate(vals, 1) if v is not None and str(v).strip()]
    
    if len(non_empty) == 0:
        continue  # 空行跳过
    
    # 判断是否数据行：第一列有值且有数值列
    has_numbers = any(v is not None and str(v).replace(',', '').replace('.', '').replace('-', '').strip().isdigit() 
                      for v in vals[2:] if v is not None)
    first_col = str(vals[0]).strip() if vals[0] else ''
    
    if 0 < len(non_empty) <= 2 and not has_numbers:
        sparse_rows.append((r, non_empty))
    elif first_col and has_numbers:
        data_row_count += 1

print(f"=== 稀疏行（续行）统计 ===")
print(f"数据行总数: {data_row_count}")
print(f"稀疏行总数: {len(sparse_rows)}")
print(f"\n稀疏行详情：")
for r, cells in sparse_rows:
    # 获取上一行同列的内容
    prev_vals = [ws.cell(r-1, c).value for c in range(1, ws.max_column + 1)]
    prev_col_vals = {c: str(prev_vals[c]).strip() if prev_vals[c] else '' for c in range(len(prev_vals))}
    continuation_info = []
    for col, text in cells:
        prev_text = prev_col_vals.get(col, '')
        continuation_info.append(f"Col{col}: '{prev_text}' + '{text}'")
    print(f"  Row {r}: {continuation_info}")

wb.close()
