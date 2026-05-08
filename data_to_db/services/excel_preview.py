"""Service: excel_preview — 读取 Excel 元信息 + 前 N 行预览数据"""

import os
import struct
import logging

import openpyxl

import services.xlrd_patch  # noqa: F401  确保 xlrd 已 patch（幂等）
from services.excel_utils import is_empty_row, is_xls_file

logger = logging.getLogger("datadeal")


def list_sheets(file_path: str) -> dict:
    """轻量接口：仅返回 sheet 名列表与文件大小。

    返回:
        {"sheet_names": list[str], "file_size": int, "is_xls": bool, "error": str?}

    相比 run() 不读取任何单元格内容，避免在仅获取 sheet 列表时遍历整张表。
    """
    file_size = os.path.getsize(file_path)
    is_xls = is_xls_file(file_path)
    if is_xls:
        return _list_sheets_xls(file_path, file_size)
    return _list_sheets_xlsx(file_path, file_size)


def _list_sheets_xlsx(file_path: str, file_size: int) -> dict:
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        # openpyxl 失败 → 回退 xlrd
        try:
            return _list_sheets_xls(file_path, file_size)
        except Exception as e2:
            return {"sheet_names": [], "file_size": file_size, "is_xls": False,
                    "error": f"文件打开失败(openpyxl: {e}; xlrd: {e2})"}
    try:
        names = list(wb.sheetnames)
    finally:
        wb.close()
    return {"sheet_names": names, "file_size": file_size, "is_xls": False}


def _list_sheets_xls(file_path: str, file_size: int) -> dict:
    import xlrd
    # xlrd 打开后 sheet_names 已经可用，不必加载格式信息
    try:
        wb = xlrd.open_workbook(file_path, formatting_info=False, on_demand=True)
        names = list(wb.sheet_names())
        return {"sheet_names": names, "file_size": file_size, "is_xls": True}
    except Exception as e:
        # xlrd 也打不开 → BIFF8 直接解析
        try:
            _, _, sheets_data, sheet_names = _read_xls_biff8(file_path)
            return {"sheet_names": list(sheet_names), "file_size": file_size, "is_xls": True}
        except Exception as e2:
            return {"sheet_names": [], "file_size": file_size, "is_xls": True,
                    "error": f"文件打开失败(xlrd: {e}; biff8: {e2})"}


def run(file_path: str, sheet_index: int = 0, preview_rows: int = 20) -> dict:
    """
    读取文件元信息 + 前N行预览数据。

    返回:
        {
            "file_path": str,
            "file_size": int,
            "sheet_names": list[str],
            "sheet_index": int,
            "sheet_name": str,
            "max_row": int,
            "max_col": int,
            "merged_count": int,
            "preview_data": list[list],  # 前N行数据
        }
    """
    file_size = os.path.getsize(file_path)
    is_xls = is_xls_file(file_path)

    if is_xls:
        return _preview_xls(file_path, sheet_index, preview_rows, file_size)
    else:
        return _preview_xlsx(file_path, sheet_index, preview_rows, file_size)


def read_first_cols(file_path: str, sheet_index: int = 0,
                    n_cols: int = 2, max_rows: int = 500) -> dict:
    """
    读取 Excel 前 n_cols 列的纵向数据（过滤尾部空行），供 LLM 判断表头/数据行位置。

    返回:
        {
            "first_col_data": list[list],   # 前 n_cols 列数据，已过滤尾部空行
            "raw_row_count": int,            # 原始总行数（含尾部空行）
        }
    """
    is_xls = is_xls_file(file_path)
    if is_xls:
        return _read_first_cols_xls(file_path, sheet_index, n_cols, max_rows)
    else:
        return _read_first_cols_xlsx(file_path, sheet_index, n_cols, max_rows)


def _read_first_cols_xlsx(file_path: str, sheet_index: int,
                           n_cols: int, max_rows: int) -> dict:
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception:
        # openpyxl 解析失败，尝试用 xlrd 回退
        try:
            return _read_first_cols_xls(file_path, sheet_index, n_cols, max_rows)
        except Exception:
            return {"first_col_data": [], "raw_row_count": 0}

    sheet_names = wb.sheetnames
    if sheet_index >= len(sheet_names):
        wb.close()
        return {"first_col_data": [], "raw_row_count": 0}

    ws = wb[sheet_names[sheet_index]]
    raw_row_count = ws.max_row or 0

    # 填充合并单元格（只处理前 n_cols 列）
    merged = {}
    for merge_range in ws.merged_cells.ranges:
        if merge_range.min_col > n_cols:
            continue
        val = ws.cell(merge_range.min_row, merge_range.min_col).value
        for r in range(merge_range.min_row, merge_range.max_row + 1):
            for c in range(merge_range.min_col, min(merge_range.max_col + 1, n_cols + 1)):
                merged[(r, c)] = val

    rows = []
    for row in ws.iter_rows(min_row=1, max_row=min(max_rows, ws.max_row or 1),
                            min_col=1, max_col=n_cols):
        row_data = [merged.get((cell.row, cell.column), cell.value) for cell in row]
        rows.append(row_data)

    wb.close()
    # 过滤尾部空行
    while rows and is_empty_row(rows[-1]):
        rows.pop()
    return {"first_col_data": rows, "raw_row_count": raw_row_count}


def _read_first_cols_xls(file_path: str, sheet_index: int,
                          n_cols: int, max_rows: int) -> dict:
    import xlrd
    try:
        wb = xlrd.open_workbook(file_path, formatting_info=True)
    except Exception:
        # xlrd 失败，尝试 BIFF8 直接解析回退
        logger.warning(f"xlrd 打开 .xls 失败，尝试 BIFF8 回退: {file_path}")
        try:
            return _read_first_cols_xls_via_biff8(file_path, sheet_index, n_cols, max_rows)
        except Exception:
            return {"first_col_data": [], "raw_row_count": 0}
    sheet_names = wb.sheet_names()
    if sheet_index >= len(sheet_names):
        return {"first_col_data": [], "raw_row_count": 0}

    ws = wb.sheet_by_index(sheet_index)
    raw_row_count = ws.nrows
    actual_n_cols = min(n_cols, ws.ncols)

    # 填充合并单元格
    # xlrd 的 merged_cells 返回 tuple 列表: (min_row, max_row, min_col, max_col)
    merged = {}
    for cr in ws.merged_cells:
        cr_min_row, cr_max_row, cr_min_col, cr_max_col = cr[0], cr[1], cr[2], cr[3]
        if cr_min_col >= actual_n_cols:
            continue
        val = ws.cell_value(cr_min_row, cr_min_col)
        for r in range(cr_min_row, cr_max_row):
            for c in range(cr_min_col, min(cr_max_col, actual_n_cols)):
                merged[(r, c)] = val

    rows = []
    for r in range(min(max_rows, ws.nrows)):
        row_data = [merged.get((r, c), ws.cell_value(r, c)) for c in range(actual_n_cols)]
        rows.append(row_data)

    while rows and is_empty_row(rows[-1]):
        rows.pop()
    return {"first_col_data": rows, "raw_row_count": raw_row_count}


def _preview_xlsx(file_path: str, sheet_index: int, preview_rows: int, file_size: int) -> dict:
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        # openpyxl 解析失败，尝试用 xlrd 回退
        logger.warning(f"openpyxl 解析失败({e})，尝试 xlrd 回退: {file_path}")
        try:
            return _preview_xls(file_path, sheet_index, preview_rows, file_size)
        except Exception as e2:
            return {"error": f"文件打开失败(openpyxl: {e}; xlrd: {e2})"}

    sheet_names = wb.sheetnames

    if sheet_index >= len(sheet_names):
        wb.close()
        return {"error": f"sheet_index {sheet_index} out of range ({len(sheet_names)} sheets)"}

    ws = wb[sheet_names[sheet_index]]
    merged_count = len(ws.merged_cells.ranges)

    # 填充合并单元格
    merged = {}
    for merge_range in ws.merged_cells.ranges:
        val = ws.cell(merge_range.min_row, merge_range.min_col).value
        for r in range(merge_range.min_row, merge_range.max_row + 1):
            for c in range(merge_range.min_col, merge_range.max_col + 1):
                merged[(r, c)] = val

    preview_data = []
    for row in ws.iter_rows(min_row=1, max_row=min(preview_rows, ws.max_row or 1),
                            max_col=ws.max_column or 1):
        row_data = []
        for cell in row:
            val = merged.get((cell.row, cell.column), cell.value)
            row_data.append(val)
        preview_data.append(row_data)

    result = {
        "file_path": file_path,
        "file_size": file_size,
        "sheet_names": sheet_names,
        "sheet_index": sheet_index,
        "sheet_name": sheet_names[sheet_index],
        "max_row": ws.max_row or 0,
        "max_col": ws.max_column or 0,
        "merged_count": merged_count,
        "preview_data": preview_data,
        "is_xls": False,
    }
    wb.close()
    return result


def _preview_xls(file_path: str, sheet_index: int, preview_rows: int, file_size: int) -> dict:
    import xlrd
    try:
        wb = xlrd.open_workbook(file_path, formatting_info=True)
    except Exception as e:
        # xlrd 不支持 formatting_info 时尝试不带格式打开
        try:
            wb = xlrd.open_workbook(file_path, formatting_info=False)
        except Exception as e2:
            # xlrd 均失败，尝试 BIFF8 直接解析回退
            logger.warning(f"xlrd 打开 .xls 失败({e2})，尝试 BIFF8 回退: {file_path}")
            try:
                return _preview_xls_via_biff8(file_path, sheet_index, preview_rows, file_size)
            except Exception as e3:
                return {"error": f"文件打开失败(xlrd: {e2}; biff8: {e3})"}
    sheet_names = wb.sheet_names()

    if sheet_index >= len(sheet_names):
        return {"error": f"sheet_index {sheet_index} out of range ({len(sheet_names)} sheets)"}

    ws = wb.sheet_by_index(sheet_index)
    merged_count = len(ws.merged_cells)

    # 填充合并单元格
    # xlrd 的 merged_cells 返回 tuple 列表: (min_row, max_row, min_col, max_col)
    merged = {}
    for cr in ws.merged_cells:
        cr_min_row, cr_max_row, cr_min_col, cr_max_col = cr[0], cr[1], cr[2], cr[3]
        val = ws.cell_value(cr_min_row, cr_min_col)
        for r in range(cr_min_row, cr_max_row):
            for c in range(cr_min_col, cr_max_col):
                merged[(r, c)] = val

    preview_data = []
    for r in range(min(preview_rows, ws.nrows)):
        row_data = []
        for c in range(ws.ncols):
            val = merged.get((r, c), ws.cell_value(r, c))
            row_data.append(val)
        preview_data.append(row_data)

    return {
        "file_path": file_path,
        "file_size": file_size,
        "sheet_names": list(sheet_names),
        "sheet_index": sheet_index,
        "sheet_name": sheet_names[sheet_index],
        "max_row": ws.nrows,
        "max_col": ws.ncols,
        "merged_count": merged_count,
        "preview_data": preview_data,
        "is_xls": True,
    }


# ============================================================================
# BIFF8 回退解析器 — 当 xlrd 无法打开 .xls 文件时，直接从 OLE2 提取 Workbook 流解析
# ============================================================================

def _decode_rk(rk_val: int):
    """Decode an RK encoded number (BIFF8 format)"""
    is_int = rk_val & 0x02
    div_100 = rk_val & 0x01
    if is_int:
        int_val = struct.unpack('<i', struct.pack('<I', rk_val & 0xFFFFFFFC))[0] >> 2
        val = int_val / 100.0 if div_100 else int_val
    else:
        rk_bytes = struct.pack('<Q', rk_val & 0xFFFFFFFFFC)
        val = struct.unpack('<d', rk_bytes)[0]
        if div_100:
            val /= 100.0
    if isinstance(val, float) and val == int(val):
        val = int(val)
    return val


def _parse_biff8_sst(rdata: bytes) -> list:
    """解析 BIFF8 SST (Shared String Table) record"""
    sst = []
    unique = struct.unpack_from('<I', rdata, 4)[0]
    pos = 8
    for _ in range(unique):
        if pos + 3 > len(rdata):
            break
        str_len = struct.unpack_from('<H', rdata, pos)[0]
        pos += 2
        flags = rdata[pos]
        pos += 1
        is_unicode = flags & 0x01
        has_rich = flags & 0x08
        has_ext = flags & 0x04

        rich_runs = 0
        if has_rich:
            rich_runs = struct.unpack_from('<H', rdata, pos)[0]
            pos += 2
        ext_size = 0
        if has_ext:
            ext_size = struct.unpack_from('<I', rdata, pos)[0]
            pos += 4

        byte_len = str_len * 2 if is_unicode else str_len
        byte_len = min(byte_len, len(rdata) - pos)
        str_bytes = rdata[pos:pos + byte_len]
        if is_unicode:
            s = str_bytes.decode('utf-16-le', errors='replace')
        else:
            s = str_bytes.decode('latin1', errors='replace')
        sst.append(s)
        pos += byte_len
        if has_rich:
            pos += rich_runs * 4
        if has_ext:
            pos += ext_size
    return sst


def _parse_biff8_boundsheet(rdata: bytes) -> str:
    """解析 BIFF8 BOUNDSHEET (0x0085) 记录中的 sheet 名。"""
    if len(rdata) < 7:
        return ""
    # offset(4) + grbit(1) + cch(1) + grbitChr(1) + bytes
    str_len = rdata[4 + 2]      # cch (sheet name 字符数)
    flag = rdata[4 + 2 + 1]     # grbitChr
    is_unicode = flag & 0x01
    name_bytes = rdata[8:]
    if is_unicode:
        return name_bytes[:str_len * 2].decode("utf-16-le", errors="replace")
    return name_bytes[:str_len].decode("latin1", errors="replace")


def _parse_biff8_workbook(wb_data: bytes) -> tuple:
    """
    从 BIFF8 Workbook 流解析所有 sheet 数据。
    返回: (sst, sheets_data, sheet_names)
        sst: 共享字符串表
        sheets_data: list of dict[(row, col) -> value]
        sheet_names: list[str] —— 与 sheets_data 顺序一致的 sheet 名
    """
    sst = []
    sheets_data = []
    sheet_names: list[str] = []
    pending_names: list[str] = []  # BOUNDSHEET 出现在 BOF(worksheet) 之前
    current_sheet = None

    pos = 0
    while pos < len(wb_data):
        if pos + 4 > len(wb_data):
            break
        rtype = struct.unpack_from('<H', wb_data, pos)[0]
        rlen = struct.unpack_from('<H', wb_data, pos + 2)[0]
        pos += 4
        rdata = wb_data[pos:pos + rlen]
        pos += rlen

        if rtype == 0x00FC:  # SST
            sst = _parse_biff8_sst(rdata)
        elif rtype == 0x0085:  # BOUNDSHEET
            name = _parse_biff8_boundsheet(rdata)
            if name:
                pending_names.append(name)
        elif rtype == 0x0809:  # BOF
            if len(rdata) >= 4:
                data_type = struct.unpack_from('<H', rdata, 2)[0]
                if data_type == 0x0010:  # worksheet
                    current_sheet = {}
                    sheets_data.append(current_sheet)
                    # 用 BOUNDSHEET 给出的对应 sheet 名（同序）
                    if len(sheet_names) < len(pending_names):
                        sheet_names.append(pending_names[len(sheet_names)])
                    else:
                        sheet_names.append(f"Sheet{len(sheet_names) + 1}")
                else:
                    current_sheet = None
        elif rtype == 0x000A:  # EOF
            current_sheet = None
        elif current_sheet is not None:
            if rtype == 0x00FD:  # LABELSST
                row = struct.unpack_from('<H', rdata, 0)[0]
                col = struct.unpack_from('<H', rdata, 2)[0]
                idx = struct.unpack_from('<I', rdata, 6)[0]
                current_sheet[(row, col)] = sst[idx] if idx < len(sst) else ""
            elif rtype == 0x0203:  # NUMBER
                row = struct.unpack_from('<H', rdata, 0)[0]
                col = struct.unpack_from('<H', rdata, 2)[0]
                val = struct.unpack_from('<d', rdata, 6)[0]
                if val == int(val):
                    val = int(val)
                current_sheet[(row, col)] = val
            elif rtype == 0x0201:  # BLANK
                row = struct.unpack_from('<H', rdata, 0)[0]
                col = struct.unpack_from('<H', rdata, 2)[0]
                current_sheet[(row, col)] = None
            elif rtype == 0x027E:  # RK
                row = struct.unpack_from('<H', rdata, 0)[0]
                col = struct.unpack_from('<H', rdata, 2)[0]
                rk_val = struct.unpack_from('<I', rdata, 6)[0]
                current_sheet[(row, col)] = _decode_rk(rk_val)
            elif rtype == 0x00BD:  # MULRK
                row = struct.unpack_from('<H', rdata, 0)[0]
                first_col = struct.unpack_from('<H', rdata, 2)[0]
                col_idx = first_col
                offset = 4
                while offset + 6 <= len(rdata) - 2:
                    rk_val = struct.unpack_from('<I', rdata, offset + 2)[0]
                    current_sheet[(row, col_idx)] = _decode_rk(rk_val)
                    col_idx += 1
                    offset += 6
            elif rtype == 0x0204:  # LABEL (BIFF2-7)
                row = struct.unpack_from('<H', rdata, 0)[0]
                col = struct.unpack_from('<H', rdata, 2)[0]
                str_len = struct.unpack_from('<H', rdata, 6)[0]
                current_sheet[(row, col)] = rdata[8:8 + str_len].decode('latin1', errors='replace')

    # 兜底补齐 sheet_names（极端情况下 BOUNDSHEET 缺失）
    while len(sheet_names) < len(sheets_data):
        sheet_names.append(f"Sheet{len(sheet_names) + 1}")
    return sst, sheets_data, sheet_names


def _sheet_dict_to_rows(sheet_data: dict, max_rows: int = 0, max_cols: int = 0) -> list:
    """将 {(row,col): value} dict 转为二维列表"""
    if not sheet_data:
        return []
    actual_max_row = max(r for r, c in sheet_data.keys())
    actual_max_col = max(c for r, c in sheet_data.keys())
    n_rows = min(actual_max_row + 1, max_rows) if max_rows else actual_max_row + 1
    n_cols = min(actual_max_col + 1, max_cols) if max_cols else actual_max_col + 1
    rows = []
    for r in range(n_rows):
        row = [sheet_data.get((r, c)) for c in range(n_cols)]
        rows.append(row)
    return rows


def _read_xls_biff8(file_path: str) -> tuple:
    """
    用 olefile + BIFF8 解析读取 .xls 文件。
    返回: (wb_data, sst, sheets_data, sheet_names) 或抛出异常。
    """
    import olefile
    ole = olefile.OleFileIO(file_path)
    if not ole.exists('Workbook') and not ole.exists('BOOK'):
        ole.close()
        raise ValueError("OLE2 文件中未找到 Workbook/BOOK 流")
    stream_name = 'Workbook' if ole.exists('Workbook') else 'BOOK'
    wb_data = ole.openstream(stream_name).read()
    ole.close()
    sst, sheets_data, sheet_names = _parse_biff8_workbook(wb_data)
    return wb_data, sst, sheets_data, sheet_names


def _preview_xls_via_biff8(file_path: str, sheet_index: int,
                             preview_rows: int, file_size: int) -> dict:
    """BIFF8 回退：预览 .xls 文件"""
    _, sst, sheets_data, sheet_names = _read_xls_biff8(file_path)

    if sheet_index >= len(sheets_data):
        return {"error": f"sheet_index {sheet_index} out of range ({len(sheets_data)} sheets)"}

    sheet_data = sheets_data[sheet_index]
    preview = _sheet_dict_to_rows(sheet_data, max_rows=preview_rows)

    max_row = max(r for r, c in sheet_data.keys()) + 1 if sheet_data else 0
    max_col = max(c for r, c in sheet_data.keys()) + 1 if sheet_data else 0

    return {
        "file_path": file_path,
        "file_size": file_size,
        "sheet_names": list(sheet_names),
        "sheet_index": sheet_index,
        "sheet_name": sheet_names[sheet_index] if sheet_index < len(sheet_names) else f"Sheet{sheet_index + 1}",
        "max_row": max_row,
        "max_col": max_col,
        "merged_count": 0,  # BIFF8 回退不解析合并单元格
        "preview_data": preview,
        "is_xls": True,
        "_biff8_fallback": True,
    }


def _read_first_cols_xls_via_biff8(file_path: str, sheet_index: int,
                                     n_cols: int, max_rows: int) -> dict:
    """BIFF8 回退：读取前 N 列纵向数据"""
    _, _, sheets_data, _ = _read_xls_biff8(file_path)

    if sheet_index >= len(sheets_data):
        return {"first_col_data": [], "raw_row_count": 0}

    sheet_data = sheets_data[sheet_index]
    max_row = max(r for r, c in sheet_data.keys()) + 1 if sheet_data else 0
    rows = _sheet_dict_to_rows(sheet_data, max_rows=max_rows, max_cols=n_cols)

    # 过滤尾部空行
    while rows and is_empty_row(rows[-1]):
        rows.pop()
    return {"first_col_data": rows, "raw_row_count": max_row}

    