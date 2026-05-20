"""Border information reading and analysis for Excel files.

Border info is a HIGH-PRIORITY structural signal:
- Solid bottom borders (ratio>=0.7) → header/data boundary or subtable end
- Dash bottom borders → vertical subtable separators (very strong signal)
- Dash right borders → horizontal split columns
- Solid borders everywhere → just formatting, skip structural meaning
"""

from __future__ import annotations

import openpyxl
from collections import Counter
from services.excel_utils import is_xls_file

DASH_STYLES = frozenset({
    'dashDot', 'dashDotDot', 'dashed', 'dotted', 'slantDashDot',
    'mediumDashed', 'mediumDashDot', 'mediumDashDotDot',
})
SOLID_STYLES = frozenset({'thin', 'medium', 'thick', 'double'})


def read_border_info(file_path: str, sheet_name) -> dict | None:
    """
    Read comprehensive border information for a sheet.

    Returns:
        {
            "rows": list[dict],  # per-row border info (0-based index into trimmed data)
            "cols": list[dict],  # per-column border info (0-based column index)
        }
        Each row dict: {
            "bottom_style": str|None, "bottom_ratio": float,
            "top_style": str|None, "top_ratio": float,
            "bottom_solid": bool, "bottom_dash": bool,
        }
        Each col dict: {
            "right_style": str|None, "right_ratio": float,
            "left_style": str|None, "left_ratio": float,
            "right_dash": bool,
        }
        Returns None if reading fails.
    """
    is_xls = is_xls_file(file_path)
    if is_xls:
        return _read_xls(file_path, sheet_name)
    return _read_xlsx(file_path, sheet_name)


def _make_row_info(bottom_styles, top_styles, n_cells):
    bs = Counter(bottom_styles).most_common(1)[0][0] if bottom_styles else None
    ts = Counter(top_styles).most_common(1)[0][0] if top_styles else None
    br = len(bottom_styles) / n_cells if n_cells > 0 else 0
    tr = len(top_styles) / n_cells if n_cells > 0 else 0
    return {
        "bottom_style": bs, "top_style": ts,
        "bottom_ratio": br, "top_ratio": tr,
        "bottom_solid": bs in SOLID_STYLES if bs else False,
        "bottom_dash": bs in DASH_STYLES if bs else False,
    }


def _make_col_info(right_styles, left_styles, n_rows):
    rs = Counter(right_styles).most_common(1)[0][0] if right_styles else None
    ls = Counter(left_styles).most_common(1)[0][0] if left_styles else None
    rr = len(right_styles) / n_rows if n_rows > 0 else 0
    lr = len(left_styles) / n_rows if n_rows > 0 else 0
    return {
        "right_style": rs, "left_style": ls,
        "right_ratio": rr, "left_ratio": lr,
        "right_dash": rs in DASH_STYLES if rs else False,
        "left_dash":  ls in DASH_STYLES if ls else False,
    }


def _read_xlsx(file_path: str, sheet_name) -> dict | None:
    try:
        # data_only=True 跳过公式 parsing，加快加载速度（border 信息与 data_only 无关）
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception:
        return None
    try:
        if isinstance(sheet_name, int):
            ws = wb[wb.sheetnames[sheet_name]]
        else:
            ws = wb[sheet_name]

        max_col = 0
        scan_limit = min(200, ws.max_column or 200)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=scan_limit):
            for cell in row:
                if cell.value is not None and str(cell.value).strip() != "":
                    max_col = max(max_col, cell.column)
        while max_col == scan_limit and scan_limit < (ws.max_column or 200):
            scan_limit = min(scan_limit * 2, ws.max_column or 200)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=scan_limit):
                for cell in row:
                    if cell.value is not None and str(cell.value).strip() != "":
                        max_col = max(max_col, cell.column)
        if max_col == 0:
            wb.close()
            return None

        # Row borders
        rows = []
        all_cells = {}  # (row_idx, col_idx) -> cell, for column analysis
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=max_col)):
            bottom_styles, top_styles = [], []
            for cell in row:
                all_cells[(row_idx, cell.column - 1)] = cell
                if cell.border.bottom and cell.border.bottom.style:
                    bottom_styles.append(cell.border.bottom.style)
                if cell.border.top and cell.border.top.style:
                    top_styles.append(cell.border.top.style)
            rows.append(_make_row_info(bottom_styles, top_styles, len(row)))

        # Column borders
        n_rows = len(rows)
        cols = []
        for col_idx in range(max_col):
            right_styles, left_styles = [], []
            for row_idx in range(n_rows):
                cell = all_cells.get((row_idx, col_idx))
                if cell:
                    if cell.border.right and cell.border.right.style:
                        right_styles.append(cell.border.right.style)
                    if cell.border.left and cell.border.left.style:
                        left_styles.append(cell.border.left.style)
            cols.append(_make_col_info(right_styles, left_styles, n_rows))

        wb.close()
        return {"rows": rows, "cols": cols}
    except Exception:
        wb.close()
        return None


def _read_xls(file_path: str, sheet_name) -> dict | None:
    import xlrd
    try:
        wb = xlrd.open_workbook(file_path, formatting_info=True)
    except Exception:
        return None
    try:
        if isinstance(sheet_name, int):
            ws = wb.sheet_by_index(sheet_name)
        else:
            ws = wb.sheet_by_name(sheet_name)
        xf_list = wb.xf_list

        rows = []
        for r in range(ws.nrows):
            bottom_styles, top_styles = [], []
            for c in range(ws.ncols):
                xf = xf_list[ws.cell_xf_index(r, c)]
                bs = _xlrd_border_style(xf.border.bottom_line_style)
                if bs:
                    bottom_styles.append(bs)
                ts = _xlrd_border_style(xf.border.top_line_style)
                if ts:
                    top_styles.append(ts)
            rows.append(_make_row_info(bottom_styles, top_styles, ws.ncols))

        cols = []
        for c in range(ws.ncols):
            right_styles, left_styles = [], []
            for r in range(ws.nrows):
                xf = xf_list[ws.cell_xf_index(r, c)]
                rs = _xlrd_border_style(xf.border.right_line_style)
                if rs:
                    right_styles.append(rs)
                ls = _xlrd_border_style(xf.border.left_line_style)
                if ls:
                    left_styles.append(ls)
            cols.append(_make_col_info(right_styles, left_styles, ws.nrows))

        return {"rows": rows, "cols": cols}
    except Exception:
        return None


def _xlrd_border_style(idx: int) -> str | None:
    _MAP = {
        0: None, 1: 'thin', 2: 'medium', 3: 'dashed', 4: 'dotted',
        5: 'thick', 6: 'double', 7: 'hair', 8: 'mediumDashed',
        9: 'dashDot', 10: 'mediumDashDot', 11: 'mediumDashDotDot',
        12: 'dashDotDot', 13: 'slantDashDot',
    }
    return _MAP.get(idx)


# ──────────────────────── Analysis functions ────────────────────────

def find_boundary_rows(rows_info: list, min_ratio: float = 0.7,
                       start: int = 0, end: int = None) -> list[int]:
    """Find row indices with solid or dash bottom border at >= min_ratio."""
    end = min(end or len(rows_info), len(rows_info))
    result = []
    for i in range(start, end):
        ri = rows_info[i]
        if ri.get("bottom_solid") and ri.get("bottom_ratio", 0) >= min_ratio:
            result.append(i)
        elif ri.get("bottom_dash") and ri.get("bottom_ratio", 0) >= min_ratio:
            result.append(i)
    return result


def find_dash_right_cols(cols_info: list, min_ratio: float = 0.3) -> list[int]:
    """Find column indices (0-based) with dash right border at >= min_ratio."""
    return [i for i, ci in enumerate(cols_info)
            if ci.get("right_dash") and ci.get("right_ratio", 0) >= min_ratio]


def find_dash_left_cols(cols_info: list, min_ratio: float = 0.01) -> list[int]:
    """Find column indices (0-based) with dash left border at >= min_ratio.

    在"横向+纵向"混合表中，两段之间的纵向分割线常以下一段首列的 left=dashDot
    形式出现（而非前一段尾列的 right 边框）。min_ratio 设得较低是因为只在每个
    分页块的最后一行（约 58/2793 ≈ 2%）出现。
    """
    return [i for i, ci in enumerate(cols_info)
            if ci.get("left_dash") and ci.get("left_ratio", 0) >= min_ratio]


def detect_header_end_by_border(rows_info: list) -> int | None:
    """
    Detect header end row using border info.

    The header_end is the last row of the header region. It's typically a row
    with a bottom border, followed by many rows without bottom borders (data region).

    Rule (4): If the majority of ALL rows have bottom borders (>70%), the borders
    are just formatting and provide no structural signal — return None.
    """
    if not rows_info:
        return None

    # Rule (4): Check if borders are just formatting (too common to be meaningful)
    boundary_rows = find_boundary_rows(rows_info, min_ratio=0.7)
    if not boundary_rows:
        return None
    border_density = len(boundary_rows) / len(rows_info)
    if border_density > 0.7:
        # Too many boundary rows — borders are just formatting
        return None

    # Find the boundary row with the largest gap after it
    max_gap = 0
    header_end_candidate = -1

    for i in range(len(boundary_rows)):
        row_idx = boundary_rows[i]
        next_boundary = boundary_rows[i + 1] if i + 1 < len(boundary_rows) else len(rows_info)
        gap = next_boundary - row_idx - 1

        if gap > max_gap:
            max_gap = gap
            header_end_candidate = row_idx

    if header_end_candidate >= 0 and max_gap >= 2:
        # Check if the candidate is actually a data row (has data content but also a border)
        # In some Excel files, the first data row has a top border from the header border
        # The real header_end is the row just before the data starts
        next_row_idx = header_end_candidate + 1
        if next_row_idx < len(rows_info):
            next_ri = rows_info[next_row_idx]
            # If the row after the candidate has a strong top border, the candidate
            # might be a header row that happens to not have a bottom border,
            # and the actual data starts at next_row_idx
            # → header_end should be header_end_candidate
            # But if header_end_candidate has a bottom border AND the next row also has
            # a bottom border, and there's no top border transition, then it's fine
            pass

        return header_end_candidate

    # Fallback: find the first boundary row followed by 2+ non-boundary rows
    for i in range(len(boundary_rows)):
        row_idx = boundary_rows[i]
        non_boundary = 0
        for j in range(row_idx + 1, min(row_idx + 5, len(rows_info))):
            if not _is_boundary_row(rows_info[j]):
                non_boundary += 1
        if non_boundary >= 2:
            return row_idx

    return None


def _is_boundary_row(ri: dict) -> bool:
    """Check if a row info dict represents a boundary row."""
    return (ri.get("bottom_solid") and ri.get("bottom_ratio", 0) >= 0.7) or \
           (ri.get("bottom_dash") and ri.get("bottom_ratio", 0) >= 0.3)


def detect_vertical_splits_by_border(rows_info: list, header_end: int) -> list[int]:
    """
    Detect vertical subtable split rows AFTER the header region.

    A split is a boundary row in the data region where the data region ends
    and is followed by a gap (empty/transition rows) before the next subtable.

    Key: a split row must be followed by a significant transition zone
    (multiple non-boundary rows) before the next boundary row. If two
    boundary rows are close together, they're part of the same structure
    (e.g., header rows of the next subtable), not a split.
    """
    boundary_rows = find_boundary_rows(rows_info, min_ratio=0.7, start=header_end + 1)
    if len(boundary_rows) < 2:
        dash_rows = [i for i in range(header_end + 1, len(rows_info))
                     if rows_info[i].get("bottom_dash") and rows_info[i].get("bottom_ratio", 0) >= 0.3]
        if not dash_rows:
            return []
        boundary_rows = sorted(set(dash_rows + boundary_rows))

    # A split is a boundary row where there's a "transition zone" after it:
    # at least 1 empty row (all cells empty/border-only) followed by content rows
    # (new subtable title/header). This distinguishes from header region boundaries
    # where there are no empty rows between header end and data start.
    splits = []
    for i in range(len(boundary_rows) - 1):
        row_idx = boundary_rows[i]
        next_boundary = boundary_rows[i + 1]

        # Check if there's an empty row in the gap (transition zone)
        has_empty_in_gap = False
        non_boundary_in_gap = 0
        for j in range(row_idx + 1, next_boundary):
            if not _is_boundary_row(rows_info[j]):
                non_boundary_in_gap += 1
            # Check for empty row: no bottom border signal AND likely empty
            if rows_info[j].get("bottom_ratio", 0) < 0.1 and rows_info[j].get("top_ratio", 0) < 0.1:
                has_empty_in_gap = True

        # A split needs: empty row(s) in gap + non-boundary rows (title/header)
        if has_empty_in_gap and non_boundary_in_gap >= 2:
            splits.append(row_idx)

    return splits


def _classify_by_border_info(rows_info: list, cols_info: list) -> dict | None:
    """纯逻辑层：根据已解析的 rows_info / cols_info 判断策略。

    与 pre_classify_by_border 的区别：不读文件，便于单元测试和复用。
    """
    # ── 规则 0: 同时存在纵向点划左线 + 水平实线表头 → 混合表 ──────────────────
    dash_left_cols = find_dash_left_cols(cols_info, min_ratio=0.01)
    header_end_candidate = detect_header_end_by_border(rows_info)
    if dash_left_cols and header_end_candidate is not None:
        return {
            "strategy": "strategy_horizontal_vertical",
            "confidence": 0.95,
            "params": {"split_col": dash_left_cols[0]},
            "source": "border_preclassify",
            "border_detail": (
                f"left_dash_cols={dash_left_cols}, "
                f"header_end={header_end_candidate}"
            ),
        }

    # ── 规则 1: 横向点划底线 → 纵向多子表 ──────────────────────────────────────
    # Only count dashes in the "data region" — skip the first few rows (likely header)
    # A dash row is meaningful if it appears after at least 2 non-dash rows
    non_dash_streak = 0
    meaningful_dash_rows = []
    for i, ri in enumerate(rows_info):
        if ri.get("bottom_dash") and ri.get("bottom_ratio", 0) >= 0.3:
            if non_dash_streak >= 2:
                meaningful_dash_rows.append(i)
            non_dash_streak = 0
        elif ri.get("bottom_solid") and ri.get("bottom_ratio", 0) >= 0.7:
            non_dash_streak = 0  # solid border resets streak too
        else:
            non_dash_streak += 1

    if len(meaningful_dash_rows) >= 1:
        return {
            "strategy": "strategy_vertical_subtable",
            "confidence": 0.85,
            "params": {"border_dash_rows": meaningful_dash_rows},
            "source": "border_preclassify",
        }

    # ── 规则 2: 纵向点划右线 → 横向分区 ────────────────────────────────────────
    dash_right_cols = find_dash_right_cols(cols_info, min_ratio=0.3)
    if dash_right_cols:
        return {
            "strategy": "strategy_horizontal_split",
            "confidence": 0.85,
            "params": {"border_split_cols": dash_right_cols},
            "source": "border_preclassify",
        }

    return None


def pre_classify_by_border(file_path: str, sheet_name) -> dict | None:
    """
    Pre-classify a sheet based on border signals, before calling LLM Classifier.

    Returns:
        {
            "strategy": "strategy_vertical_subtable" | "strategy_horizontal_split"
                      | "strategy_horizontal_vertical",
            "confidence": float,
            "params": {},
            "source": "border_preclassify"
        }
        or None if no clear border signal is found.
    """
    border_info = read_border_info(file_path, sheet_name)
    if not border_info:
        return None

    rows_info = border_info.get("rows", [])
    cols_info = border_info.get("cols", [])
    return _classify_by_border_info(rows_info, cols_info)



def detect_horizontal_split_cols(cols_info: list, rows_info: list) -> list[int]:
    """
    Detect horizontal split columns.

    Signals:
    1. Dash right borders on some columns (very strong signal)
    2. Gap regions where right_border_ratio drops vs surrounding data columns
    """
    result = []

    # Signal 1: Dash right borders
    dash_cols = find_dash_right_cols(cols_info)
    result.extend(dash_cols)

    # Signal 2: Find gap regions (consecutive low-ratio columns between high-ratio columns)
    if len(cols_info) >= 4:
        ratios = [ci.get("right_ratio", 0) for ci in cols_info]
        # Find segments of low-ratio columns
        in_gap = False
        gap_start = -1
        for i in range(len(ratios)):
            is_low = ratios[i] < 0.3
            if is_low and not in_gap:
                gap_start = i
                in_gap = True
            elif not is_low and in_gap:
                # End of gap — check if it's between high-ratio columns
                # Previous column before gap should be high ratio
                prev_high = gap_start > 0 and ratios[gap_start - 1] > 0.5
                # Current column (after gap) should be high ratio
                next_high = ratios[i] > 0.5
                if prev_high and next_high:
                    # This is a gap between data regions → split point
                    result.append(gap_start)
                in_gap = False
        # Handle gap at the end
        if in_gap and gap_start > 0 and ratios[gap_start - 1] > 0.5:
            result.append(gap_start)

    return sorted(set(result))

    