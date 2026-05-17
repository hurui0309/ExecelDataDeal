"""Service: excel_reader — 统一 Excel 读取（含合并单元格填充、可选框线检测）"""

import copy
import logging
import os
import threading

import openpyxl
from services.excel_utils import (
    is_xls_file, DASH_STYLES, SOLID_STYLES,
    BORDER_SOLID_XLRD, BORDER_DASH_XLRD,
)

logger = logging.getLogger("datadeal")


# ──────────────────────── 缓存 ────────────────────────
# 同一个 sheet 在一次任务里可能被多次读：classifier 预览之外，worker
# 还会调 strategy.run()；策略 fallback 时会再调一次。这里基于
# (file_path, sheet_name, read_border) + 文件 mtime 做轻量缓存。
#
# 内存占用：每个 sheet 几 MB；最大 32 entry 是经验值，单批运行
# 通常工作集 ≤ 一个文件的全部 sheet 数。

_CACHE_MAX = 32
_cache: "dict[tuple, tuple[float, tuple]]" = {}
_cache_lock = threading.Lock()


def _cache_get(key, mtime):
    with _cache_lock:
        entry = _cache.get(key)
        if entry is None:
            return None
        cached_mtime, value = entry
        if cached_mtime != mtime:
            # 文件已变更
            _cache.pop(key, None)
            return None
        return value


def _cache_put(key, mtime, value):
    with _cache_lock:
        if len(_cache) >= _CACHE_MAX:
            # FIFO 简单淘汰：删除任意一个最早 key（dict 保持插入序）
            try:
                _cache.pop(next(iter(_cache)))
            except StopIteration:
                pass
        _cache[key] = (mtime, value)


def clear_cache():
    """清空 read_sheet 缓存（外部测试或长时运行场景手动调用）。"""
    with _cache_lock:
        _cache.clear()


def read_sheet(file_path: str, sheet_name, read_border: bool = False):
    """
    统一读取 Excel 工作表数据（含合并单元格填充）。

    参数:
        file_path: 文件路径
        sheet_name: 工作表名或索引
        read_border: 是否读取行级水平分隔线信息

    返回:
        (data, row_has_hborder)
        - data: list[list] — 二维数据，合并单元格已填充，续行已合并
        - row_has_hborder: list[bool] — 仅当 read_border=True 时有值，否则为空列表

    实现细节：使用 (file_path, sheet_name, read_border) + 文件 mtime
    做缓存命中。返回值会做 deep copy 让调用方修改 data 不污染缓存。
    """
    try:
        mtime = os.path.getmtime(file_path)
    except OSError:
        mtime = None

    cache_key = (os.path.normcase(os.path.abspath(file_path)), sheet_name, read_border)

    if mtime is not None:
        cached = _cache_get(cache_key, mtime)
        if cached is not None:
            data, row_has_hborder = cached
            return copy.deepcopy(data), list(row_has_hborder)

    if is_xls_file(file_path):
        result = _read_xls(file_path, sheet_name, read_border)
    else:
        result = _read_xlsx(file_path, sheet_name, read_border)

    data, row_has_hborder = result

    if mtime is not None and result and result[0]:
        # 只缓存非空结果，避免读失败被一直缓存
        _cache_put(cache_key, mtime, (result[0], list(result[1] or [])))
        return copy.deepcopy(result[0]), list(result[1] or [])
    return result


def _scan_max_col(ws) -> int:
    """扫描工作表，确定实际有数据的最大列数"""
    max_col = 0
    scan_limit = min(200, ws.max_column) if ws.max_column else 200
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=scan_limit):
        for cell in row:
            if cell.value is not None and str(cell.value).strip() != "":
                max_col = max(max_col, cell.column)
    while max_col == scan_limit and scan_limit < (ws.max_column or 200):
        scan_limit = min(scan_limit * 2, ws.max_column or 200)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=scan_limit):
            for cell in row:
                if cell.value is not None and str(cell.value).strip() != "":
                    max_col = max(max_col, cell.column)
    return max_col


def _read_xlsx(file_path: str, sheet_name, read_border: bool):
    """读取 .xlsx 文件"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if isinstance(sheet_name, int):
        ws = wb[wb.sheetnames[sheet_name]]
    else:
        ws = wb[sheet_name]

    # 填充合并单元格
    merged = {}
    for merge_range in ws.merged_cells.ranges:
        val = ws.cell(merge_range.min_row, merge_range.min_col).value
        for r in range(merge_range.min_row, merge_range.max_row + 1):
            for c in range(merge_range.min_col, merge_range.max_col + 1):
                merged[(r, c)] = val

    # 扫描最大列数
    max_col = _scan_max_col(ws)
    if max_col == 0:
        wb.close()
        return [], []

    # 读取数据（可选框线）
    data = []
    row_has_hborder = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=max_col):
        row_data = [merged.get((cell.row, cell.column), cell.value) for cell in row]
        data.append(row_data)
        if read_border:
            hborder_count = 0
            for cell in row:
                if cell.border.bottom and cell.border.bottom.style:
                    if cell.border.bottom.style in SOLID_STYLES or cell.border.bottom.style in DASH_STYLES:
                        hborder_count += 1
            row_has_hborder.append(hborder_count > len(row) * 0.5)
    wb.close()
    return data, row_has_hborder


def _read_xls(file_path: str, sheet_name, read_border: bool):
    """读取 .xls 文件"""
    import xlrd
    try:
        wb = xlrd.open_workbook(file_path, formatting_info=True)
    except Exception:
        try:
            wb = xlrd.open_workbook(file_path, formatting_info=False)
            read_border = False  # 无格式信息时无法读取边框
        except Exception:
            return _read_xls_via_biff8(file_path, sheet_name), []

    if isinstance(sheet_name, int):
        ws = wb.sheet_by_index(sheet_name)
    else:
        ws = wb.sheet_by_name(sheet_name)

    # 填充合并单元格
    merged = {}
    for cr in ws.merged_cells:
        rlo, rhi, clo, chi = cr[0], cr[1], cr[2], cr[3]
        val = ws.cell_value(rlo, clo)
        for r in range(rlo, rhi):
            for c in range(clo, chi):
                merged[(r, c)] = val

    data = []
    row_has_hborder = []
    xf_list = wb.xf_list if read_border else None

    for r in range(ws.nrows):
        row_data = [merged.get((r, c), ws.cell_value(r, c)) for c in range(ws.ncols)]
        data.append(row_data)
        if read_border and xf_list:
            hborder_count = 0
            for c in range(ws.ncols):
                xf = xf_list[ws.cell_xf_index(r, c)]
                bs = xf.border.bottom_line_style
                if bs in BORDER_SOLID_XLRD or bs in BORDER_DASH_XLRD:
                    hborder_count += 1
            row_has_hborder.append(hborder_count > ws.ncols * 0.5)

    return data, row_has_hborder


def _read_xls_via_biff8(file_path: str, sheet_name):
    """BIFF8 回退：从 OLE2 提取 Workbook 流解析 .xls 文件。

    sheet_name 可以是索引（int）或 sheet 名（str），通过 BIFF8 BOUNDSHEET 解析的
    sheet_names 做名字 → 索引映射，避免回退路径永远读 sheet 0 的 bug。
    """
    from services.excel_preview import _read_xls_biff8, _sheet_dict_to_rows
    _, _, sheets_data, sheet_names = _read_xls_biff8(file_path)

    if isinstance(sheet_name, int):
        sheet_index = sheet_name
    else:
        try:
            sheet_index = sheet_names.index(sheet_name)
        except ValueError:
            # 名字找不到，安全起见返回空（caller 会感知并按 ERROR 走）
            return []

    if sheet_index >= len(sheets_data):
        return []

    sheet_data = sheets_data[sheet_index]
    return _sheet_dict_to_rows(sheet_data)

    